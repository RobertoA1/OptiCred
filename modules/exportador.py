# modules/exportador.py
"""
Módulo para exportar resultados a Excel y PDF
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

def exportar_a_excel(df_resultados, creditos, sistema, tipo_credito):
    """
    Exporta la comparación de créditos a un archivo Excel con formato profesional.
    
    Args:
        df_resultados: DataFrame con los resultados de la comparación
        creditos: Lista de diccionarios con los datos de cada crédito
        sistema: 'frances' o 'aleman'
        tipo_credito: Tipo de crédito (Consumo, Hipotecario, etc.)
    
    Returns:
        BytesIO object con el archivo Excel
    """
    from modules.amortizacion import generar_tabla_francesa, generar_tabla_alemana
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # ===== HOJA 1: RESUMEN COMPARATIVO =====
    ws_resumen = wb.create_sheet("Resumen Comparativo", 0)
    
    # Título
    ws_resumen['A1'] = 'COMPARACIÓN DE CRÉDITOS - OPTICRED'
    ws_resumen['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_resumen['A1'].fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A1:H1')
    
    # Información general
    ws_resumen['A3'] = 'Sistema de Amortización:'
    ws_resumen['B3'] = sistema.capitalize()
    ws_resumen['A4'] = 'Tipo de Crédito:'
    ws_resumen['B4'] = tipo_credito
    ws_resumen['A5'] = 'Número de opciones:'
    ws_resumen['B5'] = len(df_resultados)
    
    # Hacer negrita
    for cell in ['A3', 'A4', 'A5']:
        ws_resumen[cell].font = Font(bold=True)
    
    # Tabla de resultados (comenzando en fila 7)
    row_start = 7
    
    # Encabezados
    headers = list(df_resultados.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=row_start, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row_num, row_data in enumerate(df_resultados.values, row_start + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_resumen.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='center' if col_num == 1 else 'right')
            
            # Formato condicional: resaltar la mejor opción
            if row_num == row_start + 1 + df_resultados['Costo Total'].idxmin():
                cell.fill = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
    
    # ✅ CORRECCIÓN: Ajustar ancho de columnas SIN acceder a MergedCell
    for column_cells in ws_resumen.columns:
        max_length = 0
        # Obtener la letra de columna del primer elemento NO MERGED
        column_letter = None
        
        for cell in column_cells:
            # Saltar celdas combinadas
            if cell.coordinate in ws_resumen.merged_cells:
                continue
            
            # Obtener letra de columna
            if column_letter is None:
                column_letter = cell.column_letter
            
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Ajustar ancho si encontramos la columna
        if column_letter:
            adjusted_width = min(max_length + 2, 25)
            ws_resumen.column_dimensions[column_letter].width = adjusted_width
    
    # ===== HOJAS 2-N: TABLA DE AMORTIZACIÓN DE CADA CRÉDITO =====
    for i, credito in enumerate(creditos):
        # Generar tabla según sistema
        if sistema == 'frances':
            tabla = generar_tabla_francesa(credito['monto'], credito['tea'], credito['plazo'])
        else:
            tabla = generar_tabla_alemana(credito['monto'], credito['tea'], credito['plazo'])
        
        # Crear hoja
        ws_tabla = wb.create_sheet(f"{credito['banco'][:20]}", i + 1)
        
        # Título
        ws_tabla['A1'] = f"TABLA DE AMORTIZACIÓN - {credito['banco']}"
        ws_tabla['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_tabla['A1'].fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        ws_tabla.merge_cells('A1:F1')
        
        # Información del crédito
        ws_tabla['A3'] = 'Monto:'
        ws_tabla['B3'] = f"S/ {credito['monto']:,.2f}"
        ws_tabla['A4'] = 'TEA:'
        ws_tabla['B4'] = f"{credito['tea']*100:.2f}%"
        ws_tabla['A5'] = 'Plazo:'
        ws_tabla['B5'] = f"{credito['plazo']} meses"
        
        for cell in ['A3', 'A4', 'A5']:
            ws_tabla[cell].font = Font(bold=True)
        
        # Insertar tabla de amortización
        row_start_tabla = 7
        
        # Encabezados
        for col_num, header in enumerate(tabla.columns, 1):
            cell = ws_tabla.cell(row=row_start_tabla, column=col_num)
            cell.value = header.replace('_', ' ').title()
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, row_data in enumerate(tabla.values, row_start_tabla + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_tabla.cell(row=row_num, column=col_num)
                
                if col_num == 1:  # Columna de mes
                    cell.value = int(value)
                    cell.number_format = '0'
                else:  # Columnas de dinero
                    cell.value = float(value)
                    cell.number_format = '"S/ "#,##0.00'
                
                cell.alignment = Alignment(horizontal='center' if col_num == 1 else 'right')
        
        # Totales al final
        total_row = row_start_tabla + len(tabla) + 1
        ws_tabla.cell(row=total_row, column=1).value = 'TOTAL'
        ws_tabla.cell(row=total_row, column=1).font = Font(bold=True)
        
        ws_tabla.cell(row=total_row, column=3).value = tabla['interes'].sum()
        ws_tabla.cell(row=total_row, column=3).number_format = '"S/ "#,##0.00'
        ws_tabla.cell(row=total_row, column=3).font = Font(bold=True)
        
        ws_tabla.cell(row=total_row, column=4).value = tabla['amortizacion'].sum()
        ws_tabla.cell(row=total_row, column=4).number_format = '"S/ "#,##0.00'
        ws_tabla.cell(row=total_row, column=4).font = Font(bold=True)
        
        ws_tabla.cell(row=total_row, column=5).value = tabla['cuota'].sum()
        ws_tabla.cell(row=total_row, column=5).number_format = '"S/ "#,##0.00'
        ws_tabla.cell(row=total_row, column=5).font = Font(bold=True)
        
        # ✅ CORRECCIÓN: Ajustar ancho sin MergedCell
        for column_cells in ws_tabla.columns:
            max_length = 0
            column_letter = None
            
            for cell in column_cells:
                if cell.coordinate in ws_tabla.merged_cells:
                    continue
                
                if column_letter is None:
                    column_letter = cell.column_letter
                
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws_tabla.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def exportar_a_pdf(df_resultados, creditos, sistema, tipo_credito):
    """
    Exporta la comparación de créditos a un archivo PDF optimizado.
    
    Args:
        df_resultados: DataFrame con los resultados
        creditos: Lista de diccionarios con datos
        sistema: 'frances' o 'aleman'
        tipo_credito: Tipo de crédito
    
    Returns:
        BytesIO object con el archivo PDF
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    
    # Crear PDF en memoria
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#4A5568'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Título
    elements.append(Paragraph("COMPARACIÓN DE CRÉDITOS", title_style))
    elements.append(Paragraph("OptiCred - Sistema Inteligente", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información general
    info_data = [
        ['Sistema:', sistema.capitalize()],
        ['Tipo de Crédito:', tipo_credito],
        ['Opciones:', str(len(df_resultados))],
        ['Fecha:', pd.Timestamp.now().strftime('%d/%m/%Y')]
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONT', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4A5568')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # ===== TABLA COMPARATIVA SIMPLIFICADA =====
    elements.append(Paragraph("Resumen Comparativo", heading_style))
    
    # ✅ Crear tabla simplificada solo con columnas clave
    # Normalizar valores a 2 decimales
    df_simple = df_resultados.copy()
    
    # Seleccionar solo columnas importantes
    if sistema == 'frances':
        columnas_mostrar = ['Crédito', 'TEA (%)', 'Cuota Mensual', 'Total Intereses', 'Costo Total']
    else:
        columnas_mostrar = ['Crédito', 'TEA (%)', 'Primera Cuota', 'Total Intereses', 'Costo Total']
    
    # Filtrar columnas que existan
    columnas_existentes = [col for col in columnas_mostrar if col in df_simple.columns]
    df_simple = df_simple[columnas_existentes]
    
    # Normalizar a 2 decimales los valores numéricos
    for col in df_simple.columns:
        if df_simple[col].dtype in ['float64', 'int64'] and col != 'Crédito':
            if 'TEA' in col or 'TCEA' in col:
                df_simple[col] = df_simple[col].apply(lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) else x)
            else:
                df_simple[col] = df_simple[col].apply(lambda x: f"S/ {x:,.2f}" if isinstance(x, (int, float)) else x)
    
    # Preparar datos de la tabla
    table_data = [list(df_simple.columns)]
    for row in df_simple.values:
        table_data.append([str(val) for val in row])
    
    # Calcular anchos de columna dinámicamente
    num_cols = len(df_simple.columns)
    col_width = 6.5 * inch / num_cols  # Distribuir ancho disponible
    col_widths = [col_width] * num_cols
    
    # Crear tabla
    comparison_table = Table(table_data, colWidths=col_widths)
    
    # Encontrar índice del ganador
    idx_ganador = df_resultados['Costo Total'].idxmin()
    
    # Estilo de la tabla
    table_style = [
        # Encabezados
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Resaltar mejor opción
        ('BACKGROUND', (0, idx_ganador + 1), (-1, idx_ganador + 1), colors.HexColor('#C6F6D5')),
    ]
    
    comparison_table.setStyle(TableStyle(table_style))
    elements.append(comparison_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # ===== RECOMENDACIÓN =====
    ganador = df_resultados.loc[idx_ganador]
    elements.append(Paragraph("Recomendación", heading_style))
    
    # Formatear valores a 2 decimales
    costo_total = ganador['Costo Total']
    total_intereses = ganador['Total Intereses']
    tea = ganador['TEA (%)']
    
    recomendacion_text = f"""
    <b>🏆 Mejor Opción: {ganador['Crédito']}</b><br/>
    <br/>
    • <b>Costo Total:</b> S/ {costo_total:,.2f}<br/>
    • <b>TEA:</b> {tea:.2f}%<br/>
    • <b>Total Intereses:</b> S/ {total_intereses:,.2f}<br/>
    <br/>
    Esta opción representa el menor costo total entre todas las alternativas evaluadas.
    Se recomienda revisar los términos y condiciones detallados antes de contratar.
    """
    
    elements.append(Paragraph(recomendacion_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # ===== NOTAS AL PIE =====
    nota_style = ParagraphStyle(
        'Nota',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    )
    
    nota_text = """
    <i>Nota: Los cálculos mostrados son referenciales. Se recomienda verificar con la entidad 
    financiera los costos exactos, incluyendo seguros, comisiones adicionales y condiciones específicas.</i>
    """
    elements.append(Paragraph(nota_text, nota_style))
    
    # Construir PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return pdf_buffer